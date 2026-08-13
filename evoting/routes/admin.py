import os
import io
from functools import wraps
from datetime import datetime
from flask import request, render_template, redirect, url_for, Response, flash, send_from_directory, send_file, session
from evoting.core import (app, db, VotingConfig, KRSLog, init_db, reload_blockchain,
                          get_voting_status, get_voting_window,
                          get_client_ip, is_ip_blocked, register_failed_login,
                          clear_failed_logins, MAX_LOGIN_ATTEMPTS, BLOCK_DURATION_HOURS)


def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        password = os.environ.get('ADMIN_PASSWORD', '')
        if not password:
            return Response('ADMIN_PASSWORD tidak dikonfigurasi di server.', 503)

        ip = get_client_ip()
        # IP yang sudah diblokir ditangani lebih dulu oleh before_request,
        # tapi cek ulang di sini sebagai lapis pertahanan kedua.
        blocked, _ = is_ip_blocked(ip)
        if blocked:
            return Response('Akses diblokir sementara karena terlalu banyak percobaan gagal.', 403)

        auth = request.authorization
        if not auth or auth.password != password:
            just_blocked, sisa = register_failed_login(ip)
            if just_blocked:
                return Response(
                    f'Terlalu banyak percobaan gagal. IP Anda diblokir selama '
                    f'{BLOCK_DURATION_HOURS} jam.', 403)
            realm = (f'Admin KPU (sisa {sisa} percobaan sebelum diblokir '
                     f'{BLOCK_DURATION_HOURS} jam)')
            return Response('Login required', 401,
                            {'WWW-Authenticate': f'Basic realm="{realm}"'})

        # Login sukses: bersihkan hitungan gagal untuk IP ini.
        clear_failed_logins(ip)
        return f(*args, **kwargs)
    return decorated

@app.route('/', endpoint='admin_index')
@require_auth
def index():
    return redirect(url_for('admin_settings'))

@app.route('/results', endpoint='admin_results')
@require_auth
def results():
    bc = reload_blockchain()
    res = bc.get_results()
    invalid_blocks = bc.get_invalid_blocks()
    is_valid = len(invalid_blocks) == 0
    total_votes = sum(res.values())

    # Timeline untuk animasi replay: urutan suara berdasarkan timestamp
    # (genesis di index 0 dilewati). Dikompres ke 10 detik di sisi klien.
    timeline = sorted(
        ({'t': float(b.timestamp), 'vote': b.vote} for b in bc.chain[1:]),
        key=lambda x: x['t'],
    )
    return render_template('results.html', results=res, total_votes=total_votes,
                           is_valid=is_valid, invalid_blocks=invalid_blocks,
                           chain=bc.chain, is_admin=True, timeline=timeline)

@app.route('/admin/verify-krs')
@require_auth
def admin_krs_review():
    logs = KRSLog.query.order_by(KRSLog.id.desc()).all()
    return render_template('admin_krs_review.html', logs=logs, is_admin=True)

@app.route('/admin/verify-krs/export')
@require_auth
def admin_krs_export():
    """Ekspor seluruh log Review KRS ke berkas Excel (.xlsx)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    logs = KRSLog.query.order_by(KRSLog.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Review KRS'

    headers = ['No', 'Waktu', 'NPM', 'Nama', 'Kelas',
               'Hasil Scan QR', 'Status', 'Keterangan', 'File KRS']
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='3A557B')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for i, log in enumerate(logs, start=1):
        ws.append([
            i,
            log.timestamp or '',
            log.npm or '',
            log.name or '',
            log.kelas or '',
            log.qr_url or '',
            log.status or '',
            log.keterangan or '',
            log.filename or '',
        ])

    # Lebar kolom agar mudah dibaca.
    widths = [5, 20, 14, 28, 10, 40, 10, 45, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'review_krs_pemira_{stamp}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

@app.route('/krs-file/<path:filename>')
@require_auth
def serve_krs(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


# --- Database Pemilih Valid (gerbang password lapis-2) --------------------- #
# Password khusus fitur ini, terpisah dari ADMIN_PASSWORD login utama.
VOTER_DB_PASSWORD = 'admin123'


@app.route('/admin/pemilih-valid', methods=['GET', 'POST'], endpoint='admin_valid_voters')
@require_auth
def valid_voters():
    # Gerbang: minta password lapis-2 sebelum menampilkan data.
    if request.method == 'POST' and 'gate_password' in request.form:
        if request.form.get('gate_password') == VOTER_DB_PASSWORD:
            session['voter_db_unlocked'] = True
            return redirect(url_for('admin_valid_voters'))
        flash('Password salah. Akses ke Database Pemilih Valid ditolak.', 'error')
        return render_template('admin_valid_voters.html', unlocked=False, is_admin=True)

    if not session.get('voter_db_unlocked'):
        return render_template('admin_valid_voters.html', unlocked=False, is_admin=True)

    voters = KRSLog.query.filter_by(status='VALID').order_by(KRSLog.id.desc()).all()
    return render_template('admin_valid_voters.html',
                           unlocked=True, voters=voters, is_admin=True)


@app.route('/admin/pemilih-valid/lock', methods=['POST'], endpoint='admin_valid_voters_lock')
@require_auth
def valid_voters_lock():
    session.pop('voter_db_unlocked', None)
    return redirect(url_for('admin_valid_voters'))

@app.route('/admin/settings')
@require_auth
def admin_settings():
    config = VotingConfig.query.first()
    status = get_voting_status()
    start, end = get_voting_window()
    return render_template('admin_settings.html',
                           config=config,
                           voting_status=status,
                           voting_start=start,
                           voting_end=end,
                           is_admin=True)


@app.route('/admin/api/voting-window', methods=['POST'])
@require_auth
def api_set_voting_window():
    voting_start = request.form.get('voting_start', '').strip()
    voting_end = request.form.get('voting_end', '').strip()

    config = VotingConfig.query.first()
    if not config:
        config = VotingConfig()
        db.session.add(config)

    config.voting_start = voting_start if voting_start else None
    config.voting_end = voting_end if voting_end else None
    db.session.commit()

    flash('Jadwal voting berhasil diperbarui.', 'success')
    return redirect(url_for('admin_settings'))


@app.route('/admin/api/voting-window/clear', methods=['POST'])
@require_auth
def api_clear_voting_window():
    config = VotingConfig.query.first()
    if config:
        config.voting_start = None
        config.voting_end = None
        db.session.commit()
    flash('Jadwal voting dihapus. Voting kini tidak dibatasi waktu.', 'success')
    return redirect(url_for('admin_settings'))
